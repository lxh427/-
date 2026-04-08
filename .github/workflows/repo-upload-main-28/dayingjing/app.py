from __future__ import annotations

import csv
import io
import math
import time
from urllib.parse import quote

from flask import Flask, jsonify, make_response, request, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .backtest import backtest_module_status, backtest_schema_payload
from .cache import CacheStore
from .config import CACHE_DIR, ROOT_DIR, enabled_source_labels
from .http_client import HttpClient
from .profiles import get_live_profile
from .query_engine import QueryEngine, query_cache_key
from .stocks import resolve_stock_query

EXPORT_HEADERS = [
    "序号",
    "股票代码",
    "股票名称",
    "证据时间",
    "证据文本",
    "证据类型",
    "关联层级",
    "来源渠道",
    "来源站点",
    "来源文章",
    "原始链接",
]


def _pick_first_non_empty(values: list[object], default: str = "") -> str:
    # 统一处理 payload / query string 的多别名取值。
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return default


def _request_value(payload: dict[str, object], payload_keys: tuple[str, ...], arg_keys: tuple[str, ...], default: str = "") -> str:
    return _pick_first_non_empty(
        [payload.get(key) for key in payload_keys] + [request.args.get(key, "", type=str) for key in arg_keys],
        default=default,
    )


def _normalize_export_rows(rows: object) -> list[dict[str, object]]:
    if not isinstance(rows, list):
        return []

    normalized_rows: list[dict[str, object]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        normalized_rows.append({str(key): row.get(key, "") for key in row.keys()})
    return normalized_rows


def _coerce_filename(filename: object, default_name: str, suffix: str) -> str:
    raw = str(filename or "").strip()
    if not raw:
        return f"{default_name}{suffix}"

    if raw.lower().endswith(suffix):
        return raw

    stem = raw.rsplit(".", 1)[0] if "." in raw else raw
    return f"{stem}{suffix}"


def _ordered_export_headers(rows: list[dict[str, object]]) -> list[str]:
    if not rows:
        return []

    first_row_keys = list(rows[0].keys())
    headers = [header for header in EXPORT_HEADERS if header in first_row_keys]
    extra_headers = [header for header in first_row_keys if header not in headers]
    return headers + extra_headers


def _estimate_row_height(row: dict[str, object]) -> float:
    # 按多行字段的内容长度估算行高，避免导出后关键信息被截断。
    text_columns = (
        str(row.get("证据文本", "") or ""),
        str(row.get("来源渠道", "") or ""),
        str(row.get("来源站点", "") or ""),
        str(row.get("来源文章", "") or ""),
        str(row.get("原始链接", "") or ""),
    )
    line_count = 1

    for value in text_columns:
        segments = [segment.strip() for segment in value.splitlines() if segment.strip()]
        if not segments and value.strip():
            segments = [value.strip()]

        for segment in segments:
            line_count += max(1, math.ceil(len(segment) / 26))

    return float(min(max(24, line_count * 18), 240))


def _export_payload(payload: dict[str, object], default_name: str, suffix: str) -> tuple[list[dict[str, object]], str]:
    # 导出接口统一做行数据和文件名清洗。
    rows = _normalize_export_rows(payload.get("rows"))
    filename = _coerce_filename(payload.get("filename"), default_name, suffix)
    return rows, filename


def _apply_query_meta(payload: dict[str, object], *, cache_hit: bool, query_seconds: float) -> dict[str, object]:
    # 查询接口统一补充缓存命中和耗时元信息。
    meta = dict(payload.get("meta", {}))
    meta["cacheHit"] = cache_hit
    meta["querySeconds"] = query_seconds
    payload["meta"] = meta
    return payload


def create_app() -> Flask:
    # 应用只负责三件事：查询、导出、提供静态页面。
    app = Flask(__name__)
    cache_store = CacheStore(CACHE_DIR)
    http_client = HttpClient()
    query_engine = QueryEngine(cache_store=cache_store, http_client=http_client)

    @app.get("/")
    def index() -> object:
        return send_from_directory(ROOT_DIR, "index.html", max_age=0)

    @app.get("/app.js")
    def app_js() -> object:
        return send_from_directory(ROOT_DIR, "app.js", max_age=0)

    @app.get("/styles.css")
    def styles_css() -> object:
        return send_from_directory(ROOT_DIR, "styles.css", max_age=0)

    @app.get("/api/health")
    def health() -> object:
        return jsonify({"ok": True, "sources": enabled_source_labels()})

    @app.get("/api/backtest/status")
    def backtest_status() -> object:
        return jsonify({"ok": True, "module": backtest_module_status()})

    @app.get("/api/backtest/schema")
    def backtest_schema() -> object:
        return jsonify(backtest_schema_payload())

    @app.route("/api/query", methods=["GET", "POST"])
    def query() -> object:
        payload = request.get_json(silent=True) or {}
        query_text = _request_value(payload, ("stockQuery", "stock", "query"), ("stockQuery", "stock", "query", "q"))
        from_date = _request_value(payload, ("dateFrom", "start", "from"), ("dateFrom", "start", "from"))
        to_date = _request_value(payload, ("dateTo", "end", "to"), ("dateTo", "end", "to"))
        sensitivity = _request_value(payload, ("sensitivity",), ("sensitivity",), default="integrated")
        detail_level = _request_value(payload, ("detailLevel",), ("detailLevel",), default="full")

        if not query_text:
            return jsonify({"ok": False, "error": "缺少股票名称或代码"}), 400

        stock = resolve_stock_query(query_text, http_client=http_client, cache_store=cache_store)
        if not stock:
            return jsonify({"ok": False, "error": "未找到匹配股票"}), 404

        profile = get_live_profile(stock, http_client)
        cache_key = query_cache_key(profile, from_date, to_date, sensitivity, detail_level)
        cached = query_engine.get_cached_response(cache_key)
        if cached:
            return jsonify(_apply_query_meta(cached, cache_hit=True, query_seconds=0))

        timer = time.perf_counter()
        response = query_engine.build_query_response(profile, from_date, to_date, sensitivity, detail_level)
        response = _apply_query_meta(response, cache_hit=False, query_seconds=round(time.perf_counter() - timer, 1))
        query_engine.set_cached_response(cache_key, response)
        return jsonify(response)

    @app.post("/api/export/csv")
    def export_csv() -> object:
        payload = request.get_json(silent=True) or {}
        rows, filename = _export_payload(payload, "evidence", ".csv")

        if not rows:
            return jsonify({"ok": False, "error": "缺少导出数据"}), 400

        headers = _ordered_export_headers(rows)
        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer, lineterminator="\r\n")
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row.get(header, "") for header in headers])

        # UTF-8 BOM + sep 头，兼容中文 Excel 默认打开行为。
        body = f"sep=,\r\n{buffer.getvalue()}".encode("utf-8-sig")
        response = make_response(body)
        response.headers["Content-Type"] = "text/csv; charset=utf-8"
        response.headers["Content-Disposition"] = (
            f"attachment; filename=evidence.csv; filename*=UTF-8''{quote(filename)}"
        )
        return response

    @app.post("/api/export/xlsx")
    def export_xlsx() -> object:
        payload = request.get_json(silent=True) or {}
        rows, filename = _export_payload(payload, "evidence", ".xlsx")

        if not rows:
            return jsonify({"ok": False, "error": "缺少导出数据"}), 400

        headers = _ordered_export_headers(rows)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "证据导出"
        worksheet.freeze_panes = "A2"
        worksheet.append(headers)

        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        column_widths = {
            "序号": 8,
            "股票代码": 12,
            "股票名称": 14,
            "证据时间": 20,
            "证据文本": 72,
            "证据类型": 12,
            "关联层级": 12,
            "来源渠道": 20,
            "来源站点": 18,
            "来源文章": 32,
            "原始链接": 56,
        }

        for index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=index)
            cell.font = header_font
            cell.alignment = header_alignment
            worksheet.column_dimensions[cell.column_letter].width = column_widths.get(header, 18)

        worksheet.row_dimensions[1].height = 24

        for row_index, row in enumerate(rows, start=2):
            worksheet.append([row.get(header, "") for header in headers])
            worksheet.row_dimensions[row_index].height = _estimate_row_height(row)

            for column_index, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.alignment = cell_alignment
                if header == "原始链接" and cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"

        worksheet.auto_filter.ref = worksheet.dimensions

        output = io.BytesIO()
        workbook.save(output)

        response = make_response(output.getvalue())
        response.headers[
            "Content-Type"
        ] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response.headers["Content-Disposition"] = (
            f"attachment; filename=evidence.xlsx; filename*=UTF-8''{quote(filename)}"
        )
        return response

    return app
